import os
import io
import json
import traceback
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory, Response
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE_DIR, "data_store.json")
DASHBOARD_FILE = os.path.join(BASE_DIR, "dashboard.html")

app = Flask(__name__, static_folder=BASE_DIR, static_url_path="")

def now_iso():
    return datetime.now().isoformat(timespec="seconds")

def default_store():
    return {
        "meta": {"version": 1, "last_update": now_iso(), "source": "local"},
        "rows": [],
        "workbook": {}
    }

def load_store():
    if not os.path.exists(DATA_FILE):
        data = default_store()
        save_store(data)
        return data
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        data = default_store()
        save_store(data)
        return data

def save_store(data):
    data.setdefault("meta", {})
    data["meta"]["last_update"] = now_iso()
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def normalize_header(value, fallback):
    if value is None:
        return fallback
    text = str(value).strip()
    return text if text else fallback

def workbook_to_json(file_bytes):
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    workbook_json = {}
    merged_rows = []

    for sheet in wb.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            workbook_json[sheet.title] = []
            continue

        headers = [normalize_header(v, f"col_{i+1}") for i, v in enumerate(values[0])]
        sheet_rows = []

        for row in values[1:]:
            row_dict = {}
            has_value = False
            for i, cell in enumerate(row):
                key = headers[i] if i < len(headers) else f"col_{i+1}"
                if hasattr(cell, "strftime"):
                    try:
                        val = cell.strftime("%Y-%m-%d")
                    except Exception:
                        val = cell
                else:
                    val = cell
                if val not in (None, ""):
                    has_value = True
                row_dict[key] = val
            if has_value:
                sheet_rows.append(row_dict)

        workbook_json[sheet.title] = sheet_rows
        for item in sheet_rows:
            merged_item = {"_sheet": sheet.title}
            merged_item.update(item)
            merged_rows.append(merged_item)

    return {
        "meta": {
            "version": 1,
            "last_update": now_iso(),
            "source": "xlsx_upload",
            "sheets": list(workbook_json.keys())
        },
        "rows": merged_rows,
        "workbook": workbook_json
    }

def serve_dashboard():
    if os.path.exists(DASHBOARD_FILE):
        return send_from_directory(BASE_DIR, "dashboard.html")
    return Response("<h1>dashboard.html não encontrado</h1>", status=200, mimetype="text/html")

@app.get("/")
def home():
    return serve_dashboard()

@app.get("/dashboard.html")
def dashboard():
    return serve_dashboard()

@app.get("/api/health")
def api_health():
    return jsonify({
        "ok": True,
        "service": "contel-dashboard",
        "time": now_iso(),
        "dashboard_exists": os.path.exists(DASHBOARD_FILE),
        "data_exists": os.path.exists(DATA_FILE)
    })

@app.get("/api/data")
def api_data():
    return jsonify(load_store())

@app.post("/api/upload-xlsx")
def api_upload_xlsx():
    file = request.files.get("file")
    if not file:
        return jsonify({"ok": False, "error": "Nenhum arquivo enviado."}), 400

    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return jsonify({"ok": False, "error": "Envie um arquivo .xlsx válido."}), 400

    try:
        raw = file.read()
        data = workbook_to_json(raw)
        save_store(data)
        return jsonify({
            "ok": True,
            "message": "Planilha importada com sucesso.",
            "meta": data.get("meta", {}),
            "rows": len(data.get("rows", []))
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": f"Falha ao importar planilha: {e}"}), 500

@app.get("/api/reset")
def api_reset():
    data = default_store()
    save_store(data)
    return jsonify({"ok": True, "message": "Base resetada com sucesso."})

@app.route("/<path:path>")
def static_or_fallback(path):
    if path.startswith("api/"):
        return jsonify({"ok": False, "error": "Rota API não encontrada."}), 404

    file_path = os.path.join(BASE_DIR, path)
    if os.path.isfile(file_path):
        return send_from_directory(BASE_DIR, path)

    return serve_dashboard()

@app.errorhandler(500)
def internal_error(e):
    traceback.print_exc()
    return jsonify({
        "ok": False,
        "error": "Erro interno no servidor.",
        "detail": str(e)
    }), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    print(f"Iniciando servidor na porta {port}")
    print(f"BASE_DIR={BASE_DIR}")
    print(f"DASHBOARD_FILE existe? {os.path.exists(DASHBOARD_FILE)}")
    print(f"DATA_FILE existe? {os.path.exists(DATA_FILE)}")
    app.run(host="0.0.0.0", port=port, debug=False)
